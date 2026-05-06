"""CRM Import — CSV / XLSX parsing with explicit column mapping.

Two-phase flow:

  1. Preview — frontend uploads the file once. We parse the header row,
     a small sample of body rows (~10), and the total row count. The
     frontend renders a column-mapping UI letting the operator pick
     which source column maps onto each ARIA contact field. NOTHING is
     written to the database in this phase.

  2. Import — frontend re-uploads the same file along with the chosen
     mapping (a dict like `{"email": "Email Address", "name": "Full
     Name", "company": "Org"}`). We re-parse, apply the mapping row by
     row, dedupe by email-within-tenant, auto-create / link companies,
     and insert into crm_contacts.

The deliberate "upload twice" choice avoids server-side temp file
state — keeps the import endpoint stateless. For a 5k-row XLSX that's
~1MB on the wire; cheaper than building a session-based staging area.

Field set:
  email, name, phone, title, company, status, source, tags, notes

Notes are concatenated from any extra columns the operator maps to
"notes", separated by " | " so all the leftover detail still ends up
in the row instead of being dropped.

Uniformity rule (per the operator's request): nothing is auto-detected.
The mapping is fully explicit, so two CSVs with different header names
still produce identical column placement in crm_contacts.
"""
from __future__ import annotations

import csv
import io
import logging
import re
from typing import Any, Iterable, Optional

logger = logging.getLogger("aria.services.crm_import")


# ── Field set ─────────────────────────────────────────────────────────


# These are the ARIA-side canonical fields the operator can map columns
# onto. Frontend renders a dropdown per field with the source columns
# as options. Anything not in this set is silently ignored.
SUPPORTED_FIELDS: tuple[str, ...] = (
    "email",
    "name",
    "phone",
    "title",
    "company",
    "status",
    "source",
    "tags",
    "notes",
)

# How many sample rows to return in the /preview response. Enough for
# the operator to spot-check column choices without inflating the
# payload.
PREVIEW_SAMPLE_SIZE = 10

# Hard cap on import size. 10k contacts at ~500 bytes each ≈ 5MB
# payload, which Hostinger SMTP would take a week to drain at 1k/day
# anyway. If the operator has more than this they should batch.
MAX_IMPORT_ROWS = 10_000


# ── Status / source vocabulary ────────────────────────────────────────


# Valid contact statuses (mirrors CONTACT_STATUSES in the frontend).
# We coerce mapped status values onto this set, defaulting to "lead".
_VALID_STATUSES = {
    "lead", "qualified", "proposal", "negotiation",
    "won", "lost", "customer", "subscriber",
}


def _normalize_status(raw: str) -> str:
    """Coerce free-text status values onto the CRM vocabulary."""
    if not raw:
        return "lead"
    v = raw.strip().lower()
    if v in _VALID_STATUSES:
        return v
    # Common synonyms / alternative spellings
    aliases = {
        "prospect": "lead",
        "new": "lead",
        "interested": "qualified",
        "in progress": "negotiation",
        "active": "qualified",
        "client": "customer",
        "closed won": "won",
        "closed lost": "lost",
        "closed-won": "won",
        "closed-lost": "lost",
    }
    return aliases.get(v, "lead")


# ── File parsing ──────────────────────────────────────────────────────


def _detect_format(filename: str, content_type: str) -> str:
    """Return 'xlsx' or 'csv'. Defaults to csv when the signal is
    ambiguous so a wrong content-type header still falls through to a
    reasonable default."""
    name = (filename or "").lower()
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        return "xlsx"
    if name.endswith(".csv") or name.endswith(".tsv") or name.endswith(".txt"):
        return "csv"
    if "spreadsheetml" in (content_type or "").lower():
        return "xlsx"
    return "csv"


def _parse_csv_bytes(raw: bytes) -> list[list[str]]:
    """Decode CSV bytes into a list of row lists.

    Strips a UTF-8 BOM if present (Excel writes one when you "Save as
    CSV UTF-8") and falls back to latin-1 for older exports that aren't
    valid UTF-8.
    """
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = raw.decode("latin-1", errors="replace")
    # csv.reader handles quoted-cells-with-commas, embedded newlines,
    # and Windows / Unix line endings transparently.
    reader = csv.reader(io.StringIO(text))
    return [[(cell or "").strip() for cell in row] for row in reader if any((c or "").strip() for c in row)]


def _parse_xlsx_bytes(raw: bytes) -> list[list[str]]:
    """Decode an XLSX workbook into row lists from the FIRST sheet only.

    We don't pretend to support multi-sheet imports; the first sheet's
    header row is the schema. Cells with formulas resolve to the cached
    value (openpyxl read_only mode reads cached values by default).
    """
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RuntimeError(
            "openpyxl is required for XLSX imports — pip install openpyxl",
        ) from e

    wb = load_workbook(filename=io.BytesIO(raw), read_only=True, data_only=True)
    try:
        sheet = wb.active
        rows: list[list[str]] = []
        for row in sheet.iter_rows(values_only=True):
            cells = [
                (str(c).strip() if c is not None else "")
                for c in row
            ]
            if any(cells):
                rows.append(cells)
        return rows
    finally:
        wb.close()


def parse_file(raw: bytes, filename: str, content_type: str) -> list[list[str]]:
    """Public entry point — detect format and decode to row lists.

    Header row is rows[0]. Body is rows[1:]. Returns empty list when
    the file has no usable rows (so callers can short-circuit with a
    400 instead of indexing into a bad shape).
    """
    fmt = _detect_format(filename, content_type)
    if fmt == "xlsx":
        return _parse_xlsx_bytes(raw)
    return _parse_csv_bytes(raw)


# ── Row → contact dict ────────────────────────────────────────────────


_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


def _split_tags(value: str) -> list[str]:
    """Tags can come in as comma-, semicolon-, or pipe-separated. Returns
    a deduped lowercased list, preserving order. Empty string → []."""
    if not value:
        return []
    parts = re.split(r"[,;|]", value)
    seen: set[str] = set()
    out: list[str] = []
    for p in parts:
        t = p.strip().lower()
        if t and t not in seen:
            seen.add(t)
            out.append(t)
    return out


def build_contact_from_row(
    row: list[str],
    header: list[str],
    mapping: dict[str, str],
    extra_notes_columns: Optional[list[str]] = None,
) -> tuple[Optional[dict[str, Any]], Optional[str]]:
    """Translate one source row into a contact-shaped dict.

    `mapping` is {aria_field: source_column_name}. Source column names
    are matched against `header` case-insensitively to be forgiving of
    Excel auto-capitalization shenanigans.

    Returns (contact_dict, error). If the row can't be made into a
    valid contact (e.g. no name AND no email), returns (None, reason).

    `extra_notes_columns` is the set of source columns the operator
    chose to roll up into notes — typically the columns they didn't
    map onto a primary field. Joined with " | " in the final notes.
    """
    # Build lowercase index of column → cell
    idx = {(h or "").strip().lower(): i for i, h in enumerate(header)}

    def _cell(col_name: str) -> str:
        if not col_name:
            return ""
        i = idx.get(col_name.strip().lower(), -1)
        if i < 0 or i >= len(row):
            return ""
        return (row[i] or "").strip()

    out: dict[str, Any] = {}

    # Email — primary identity. Must match a basic shape; otherwise
    # blank so the de-dupe layer doesn't accidentally collapse two
    # different non-emails together.
    email_raw = _cell(mapping.get("email", ""))
    email = email_raw.lower() if _EMAIL_RE.match(email_raw or "") else ""
    if email:
        out["email"] = email

    name = _cell(mapping.get("name", ""))
    if name:
        out["name"] = name

    phone = _cell(mapping.get("phone", ""))
    if phone:
        out["phone"] = phone

    title = _cell(mapping.get("title", ""))
    if title:
        out["title"] = title

    company = _cell(mapping.get("company", ""))
    if company:
        out["_company_name"] = company  # consumed by importer; not stored directly

    status_raw = _cell(mapping.get("status", ""))
    if status_raw:
        out["status"] = _normalize_status(status_raw)
    else:
        out["status"] = "lead"

    source = _cell(mapping.get("source", "")) or "import"
    out["source"] = source[:50]

    tags_raw = _cell(mapping.get("tags", ""))
    tags = _split_tags(tags_raw)
    if tags:
        out["tags"] = tags

    # Build notes — start with the explicitly mapped notes column,
    # then append "Header: Value" for each unmapped column the operator
    # asked us to roll up. Keeps incidental data findable later.
    notes_parts: list[str] = []
    explicit_notes = _cell(mapping.get("notes", ""))
    if explicit_notes:
        notes_parts.append(explicit_notes)
    for extra_col in (extra_notes_columns or []):
        v = _cell(extra_col)
        if v:
            notes_parts.append(f"{extra_col}: {v}")
    if notes_parts:
        out["notes"] = " | ".join(notes_parts)[:2000]

    # A row is rejected when it has neither an email nor a name. Without
    # one of them we can't store anything useful.
    if not out.get("email") and not out.get("name"):
        return None, "row has no email and no name"

    return out, None


# ── Bulk import ───────────────────────────────────────────────────────


def import_contacts(
    tenant_id: str,
    rows: list[list[str]],
    mapping: dict[str, str],
    extra_notes_columns: Optional[list[str]] = None,
) -> dict[str, Any]:
    """Insert contacts in batch. Returns counts + error rows.

    De-dupe rules (per uniformity ask):
      - Within the upload: same email skipped after first occurrence.
      - Against existing crm_contacts: same tenant + same email skipped.

    Companies are auto-created if `mapping["company"]` is set and the
    company name doesn't already exist for the tenant. The contact's
    `company_id` then points at the matching row.
    """
    if not rows:
        return {"imported": 0, "skipped": 0, "errors": []}

    header = rows[0]
    body = rows[1:]
    if len(body) > MAX_IMPORT_ROWS:
        return {
            "imported": 0,
            "skipped": 0,
            "errors": [
                {"row": 0, "reason": f"file has {len(body)} rows (max {MAX_IMPORT_ROWS}); split it and re-upload"}
            ],
        }

    from backend.services.supabase import get_db
    sb = get_db()

    # Pre-load existing emails for this tenant so we can dedupe in one
    # query instead of N. Only pulls the email column.
    existing_emails: set[str] = set()
    try:
        existing = (
            sb.table("crm_contacts")
            .select("email")
            .eq("tenant_id", tenant_id)
            .execute()
        )
        for r in (existing.data or []):
            e = (r.get("email") or "").strip().lower()
            if e:
                existing_emails.add(e)
    except Exception as e:
        logger.warning("[crm-import] preload existing emails failed: %s", e)

    # Pre-load existing companies. Map normalized name → id so we can
    # link contacts without re-creating duplicates.
    existing_companies: dict[str, str] = {}
    try:
        comp_res = (
            sb.table("crm_companies")
            .select("id, name")
            .eq("tenant_id", tenant_id)
            .execute()
        )
        for r in (comp_res.data or []):
            n = (r.get("name") or "").strip().lower()
            if n and r.get("id"):
                existing_companies[n] = r["id"]
    except Exception as e:
        logger.warning("[crm-import] preload existing companies failed: %s", e)

    imported = 0
    skipped = 0
    errors: list[dict[str, Any]] = []
    seen_in_batch: set[str] = set()

    for i, row in enumerate(body, start=2):  # +1 for header, +1 for 1-indexed
        try:
            contact, err = build_contact_from_row(
                row, header, mapping, extra_notes_columns,
            )
            if err or not contact:
                errors.append({"row": i, "reason": err or "unknown"})
                continue

            # In-batch dedupe by email (only when we have one).
            email = contact.get("email", "")
            if email:
                if email in seen_in_batch:
                    skipped += 1
                    continue
                seen_in_batch.add(email)
                if email in existing_emails:
                    skipped += 1
                    continue

            # Resolve company_id (auto-create when needed).
            company_name = contact.pop("_company_name", "")
            if company_name:
                key = company_name.strip().lower()
                cid = existing_companies.get(key)
                if not cid:
                    try:
                        ins = (
                            sb.table("crm_companies")
                            .insert({"tenant_id": tenant_id, "name": company_name[:200]})
                            .execute()
                        )
                        if ins.data:
                            cid = ins.data[0]["id"]
                            existing_companies[key] = cid
                    except Exception as e:
                        logger.debug("[crm-import] company autocreate failed: %s", e)
                if cid:
                    contact["company_id"] = cid

            # Default name when only email was provided.
            if not contact.get("name"):
                contact["name"] = (email.split("@", 1)[0] if email else "Imported contact")

            row_to_insert = {"tenant_id": tenant_id, **contact}
            row_to_insert.setdefault("source", "import")
            row_to_insert.setdefault("status", "lead")

            try:
                sb.table("crm_contacts").insert(row_to_insert).execute()
                imported += 1
                if email:
                    existing_emails.add(email)
            except Exception as e:
                errors.append({"row": i, "reason": f"insert failed: {e}"})

        except Exception as e:
            errors.append({"row": i, "reason": f"unexpected: {e}"})

    return {
        "imported": imported,
        "skipped": skipped,
        "errors": errors[:50],  # cap so we don't return a 5MB error list
        "total_rows_in_file": len(body),
    }


# ── Preview helpers ───────────────────────────────────────────────────


def build_preview(rows: list[list[str]]) -> dict[str, Any]:
    """Return a small payload describing the file's shape."""
    if not rows:
        return {
            "headers": [],
            "sample_rows": [],
            "total_row_count": 0,
            "supported_fields": list(SUPPORTED_FIELDS),
        }
    header = rows[0]
    body = rows[1:]
    return {
        "headers": [(h or "").strip() for h in header],
        "sample_rows": body[:PREVIEW_SAMPLE_SIZE],
        "total_row_count": len(body),
        "supported_fields": list(SUPPORTED_FIELDS),
    }

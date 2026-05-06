"""Facebook Ads Report CSV Parser — normalizes Meta Ads Manager exports.

Handles column name variations and extracts structured metrics.
"""
from __future__ import annotations

import csv
import io
import logging
from typing import Any

logger = logging.getLogger("aria.fb_ads_parser")

# ── Column alias mapping ────────────────────────────────────────────────────────
# Maps common Facebook Ads Manager column names to normalized metric keys.
# Keys are lowercased for case-insensitive matching.

COLUMN_ALIASES: dict[str, str] = {
    # Spend
    "amount spent": "spend",
    "amount spent (usd)": "spend",
    "cost": "spend",
    "spend": "spend",
    "total spent": "spend",

    # Impressions
    "impressions": "impressions",
    "imps": "impressions",

    # Reach
    "reach": "reach",
    "people reached": "reach",

    # Clicks
    "clicks (all)": "clicks",
    "clicks": "clicks",
    "all clicks": "clicks",

    # Link clicks
    "link clicks": "link_clicks",
    "link click": "link_clicks",
    "outbound clicks": "link_clicks",
    "website clicks": "link_clicks",

    # CTR
    "ctr (all)": "ctr",
    "ctr": "ctr",
    "ctr (link click-through rate)": "ctr",
    "click-through rate": "ctr",

    # CPC
    "cpc (all)": "cpc",
    "cpc": "cpc",
    "cpc (cost per link click)": "cpc",
    "cost per click": "cpc",
    "cost per click (all)": "cpc",

    # CPM
    "cpm (cost per 1,000 impressions)": "cpm",
    "cpm": "cpm",
    "cost per 1,000 impressions": "cpm",

    # Conversions
    "results": "conversions",
    "conversions": "conversions",
    "leads": "conversions",
    "total conversions": "conversions",
    "actions": "conversions",
    "purchases": "conversions",

    # Cost per result
    "cost per result": "cost_per_result",
    "cost per conversion": "cost_per_result",
    "cost per lead": "cost_per_result",
    "cost per action": "cost_per_result",

    # Frequency
    "frequency": "frequency",
    "avg. frequency": "frequency",

    # Campaign identifiers
    "campaign name": "campaign_name",
    "campaign": "campaign_name",
    "campaign id": "campaign_id",

    # Ad set / Ad identifiers
    "ad set name": "ad_set_name",
    "adset name": "ad_set_name",
    "ad name": "ad_name",

    # Date range
    "reporting starts": "date_start",
    "reporting ends": "date_end",
    "date start": "date_start",
    "date end": "date_end",
    "day": "date_start",
    "date": "date_start",

    # Objective / Delivery
    "objective": "objective",
    "delivery": "delivery_status",
    "status": "status",

    # Additional metrics
    "video views": "video_views",
    "video plays": "video_views",
    "3-second video views": "video_views_3s",
    "thruplay": "thruplays",
    "post engagement": "post_engagement",
    "post engagements": "post_engagement",
    "page engagement": "page_engagement",
    "page likes": "page_likes",
    "roas": "roas",
    "return on ad spend": "roas",
    "purchase roas": "roas",
}

# Metrics that should be parsed as numbers
NUMERIC_METRICS = {
    "spend", "impressions", "reach", "clicks", "link_clicks",
    "ctr", "cpc", "cpm", "conversions", "cost_per_result",
    "frequency", "video_views", "video_views_3s", "thruplays",
    "post_engagement", "page_engagement", "page_likes", "roas",
}


def _parse_number(value: str) -> float | None:
    """Parse a number from a string, handling currency symbols and commas."""
    if not value or not value.strip():
        return None
    cleaned = value.strip().replace(",", "").replace("$", "").replace("€", "").replace("£", "").replace("%", "")
    try:
        return float(cleaned)
    except (ValueError, TypeError):
        return None


def _normalize_column(col: str) -> str | None:
    """Map a column name to its normalized metric key."""
    return COLUMN_ALIASES.get(col.strip().lower())


def parse_csv(content: str | bytes) -> dict[str, Any]:
    """Parse a Facebook Ads CSV export into structured metrics.

    Returns:
        {
            "success": True/False,
            "campaigns": [
                {
                    "campaign_name": "...",
                    "campaign_id": "...",
                    "metrics": { "spend": 123.45, ... },
                    "date_start": "...",
                    "date_end": "...",
                    "rows": 5,
                }
            ],
            "totals": { "spend": 500.00, ... },
            "row_count": 10,
            "error": "..." (only if failed)
        }
    """
    if isinstance(content, bytes):
        # Try UTF-8 first, then latin-1 fallback
        try:
            content = content.decode("utf-8-sig")  # handles BOM
        except UnicodeDecodeError:
            content = content.decode("latin-1")

    # Remove any leading empty lines or metadata rows (Facebook sometimes adds headers)
    lines = content.strip().split("\n")
    if not lines:
        return {"success": False, "error": "Empty file"}

    # Find the header row — skip non-header lines
    header_idx = 0
    for i, line in enumerate(lines):
        # Facebook Ads exports sometimes have metadata before the actual CSV header
        if any(alias in line.lower() for alias in ("campaign name", "impressions", "amount spent", "clicks")):
            header_idx = i
            break

    csv_content = "\n".join(lines[header_idx:])
    reader = csv.DictReader(io.StringIO(csv_content))

    if not reader.fieldnames:
        return {"success": False, "error": "Could not detect CSV columns"}

    # Build column mapping
    col_map: dict[str, str] = {}
    unmapped_cols: list[str] = []
    for col in reader.fieldnames:
        normalized = _normalize_column(col)
        if normalized:
            col_map[col] = normalized
        else:
            unmapped_cols.append(col)

    if not col_map:
        return {
            "success": False,
            "error": f"No recognized Facebook Ads columns found. Columns in file: {', '.join(reader.fieldnames[:10])}",
        }

    # Parse rows
    campaigns: dict[str, dict] = {}  # campaign_name -> aggregated data
    totals: dict[str, float] = {}
    row_count = 0

    for row in reader:
        row_count += 1
        campaign_name = ""
        campaign_id = ""
        row_metrics: dict[str, Any] = {}
        date_start = ""
        date_end = ""

        for csv_col, metric_key in col_map.items():
            value = row.get(csv_col, "")

            if metric_key == "campaign_name":
                campaign_name = value.strip()
            elif metric_key == "campaign_id":
                campaign_id = value.strip()
            elif metric_key == "date_start":
                date_start = value.strip()
            elif metric_key == "date_end":
                date_end = value.strip()
            elif metric_key in NUMERIC_METRICS:
                num = _parse_number(value)
                if num is not None:
                    row_metrics[metric_key] = num
            else:
                if value.strip():
                    row_metrics[metric_key] = value.strip()

        if not campaign_name:
            campaign_name = "Unknown Campaign"

        # Aggregate by campaign
        if campaign_name not in campaigns:
            campaigns[campaign_name] = {
                "campaign_name": campaign_name,
                "campaign_id": campaign_id,
                "metrics": {},
                "date_start": date_start,
                "date_end": date_end,
                "rows": 0,
            }

        camp = campaigns[campaign_name]
        camp["rows"] += 1

        # Update date range
        if date_start and (not camp["date_start"] or date_start < camp["date_start"]):
            camp["date_start"] = date_start
        if date_end and (not camp["date_end"] or date_end > camp["date_end"]):
            camp["date_end"] = date_end

        # Sum numeric metrics
        for key, val in row_metrics.items():
            if key in NUMERIC_METRICS and isinstance(val, (int, float)):
                camp["metrics"][key] = camp["metrics"].get(key, 0) + val
                totals[key] = totals.get(key, 0) + val

    # Calculate derived metrics per campaign
    for camp in campaigns.values():
        m = camp["metrics"]
        if m.get("clicks") and m.get("impressions"):
            m["ctr"] = round((m["clicks"] / m["impressions"]) * 100, 2)
        if m.get("spend") and m.get("clicks"):
            m["cpc"] = round(m["spend"] / m["clicks"], 2)
        if m.get("spend") and m.get("impressions"):
            m["cpm"] = round((m["spend"] / m["impressions"]) * 1000, 2)
        if m.get("spend") and m.get("conversions"):
            m["cost_per_result"] = round(m["spend"] / m["conversions"], 2)
        if m.get("impressions") and m.get("reach"):
            m["frequency"] = round(m["impressions"] / m["reach"], 2)
        # Round all numeric values
        for k, v in m.items():
            if isinstance(v, float):
                m[k] = round(v, 2)

    # Same for totals
    if totals.get("clicks") and totals.get("impressions"):
        totals["ctr"] = round((totals["clicks"] / totals["impressions"]) * 100, 2)
    if totals.get("spend") and totals.get("clicks"):
        totals["cpc"] = round(totals["spend"] / totals["clicks"], 2)
    if totals.get("spend") and totals.get("impressions"):
        totals["cpm"] = round((totals["spend"] / totals["impressions"]) * 1000, 2)
    if totals.get("spend") and totals.get("conversions"):
        totals["cost_per_result"] = round(totals["spend"] / totals["conversions"], 2)

    return {
        "success": True,
        "campaigns": list(campaigns.values()),
        "totals": {k: round(v, 2) for k, v in totals.items()},
        "row_count": row_count,
        "mapped_columns": list(col_map.values()),
        "unmapped_columns": unmapped_cols,
    }


# ── XLSX support ───────────────────────────────────────────────────────────────
#
# Meta Ads Manager exports as CSV by default, but operators frequently
# pull the data through Excel first (cleaning columns, applying filters,
# saving as .xlsx). Rather than telling them to "Save As CSV first",
# we accept .xlsx directly: read the first sheet via openpyxl, write
# back out as a CSV string, and feed it to the existing parse_csv path.
# Keeps every downstream behavior (column aliases, totals, date
# extraction, suggested-match logic) identical for both formats.


def _xlsx_to_csv_text(raw: bytes) -> str:
    """Convert the first sheet of an XLSX workbook into a CSV-formatted
    string. Cells containing commas / newlines / quotes get quoted by
    csv.writer the same way an Excel "Save As CSV" would. Empty rows
    are dropped so the parse_csv header-detection heuristic still works
    on workbooks that prefix their data with blank rows.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RuntimeError(
            "openpyxl is required for XLSX uploads — pip install openpyxl",
        ) from e

    wb = load_workbook(filename=io.BytesIO(raw), read_only=True, data_only=True)
    try:
        sheet = wb.active
        out = io.StringIO()
        writer = csv.writer(out)
        for row in sheet.iter_rows(values_only=True):
            cells = [
                ("" if c is None else str(c))
                for c in row
            ]
            if any(c.strip() for c in cells):
                writer.writerow(cells)
        return out.getvalue()
    finally:
        wb.close()


def parse_report(content: bytes, filename: str = "") -> dict[str, Any]:
    """Format-detecting wrapper around parse_csv.

    Dispatches to XLSX → CSV conversion when the filename ends in
    .xlsx / .xlsm; otherwise delegates straight to parse_csv. Returns
    the same structured payload regardless of source format so callers
    don't need to branch.
    """
    name = (filename or "").lower()
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        try:
            csv_text = _xlsx_to_csv_text(content)
        except Exception as e:
            return {
                "success": False,
                "error": f"Could not read Excel file: {e}",
            }
        return parse_csv(csv_text)
    return parse_csv(content)

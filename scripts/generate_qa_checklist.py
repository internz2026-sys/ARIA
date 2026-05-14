from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "ARIA QA Checklist"

# Styles
header_font = Font(bold=True, color="FFFFFF", size=12)
header_fill = PatternFill("solid", fgColor="1F4E78")
section_font = Font(bold=True, color="FFFFFF", size=11)
section_fill = PatternFill("solid", fgColor="2E75B6")
thin = Side(border_style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="center")

headers = ["#", "Section", "Test Case", "Steps / Expected Result", "Priority", "Status", "Tester", "Date", "Notes"]
ws.append(headers)
for col, _ in enumerate(headers, 1):
    c = ws.cell(row=1, column=col)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = border

sections = [
    ("Auth & Onboarding", [
        ("Sign up with email", "Submit email signup form -> confirmation email arrives -> account created", "P0"),
        ("Sign up with GitHub OAuth", "Click GitHub button -> authorize -> redirected back authenticated", "P0"),
        ("Login valid creds", "Enter valid email/password -> lands on /dashboard", "P0"),
        ("Login invalid creds", "Wrong password -> clear error message, no crash", "P1"),
        ("Logout clears session", "Click logout -> protected routes redirect to /login", "P0"),
        ("Onboarding flow completes", "/welcome -> /select-agents -> /connect -> /review without errors", "P0"),
        ("CEO conversational intake", "Complete 10-15 min intake -> GTM Playbook generated", "P0"),
        ("Tenant config persists", "POST /api/onboarding/save-config -> row in Supabase tenant_configs", "P0"),
        ("Resume mid-onboarding", "Refresh during step 2 -> resumes at correct step", "P1"),
    ]),
    ("Dashboard & Navigation", [
        ("Sidebar routes", "Click each sidebar link -> correct page loads", "P0"),
        ("KPI cards render", "Dashboard KPI cards show real data, not placeholders", "P1"),
        ("Org chart renders", "All 5 agents visible with status badges", "P1"),
        ("Agent click opens panel", "Click an agent -> recent work panel opens", "P2"),
        ("No console errors", "Open DevTools console on each page -> zero errors", "P1"),
        ("Page load < 3s", "Cold cache load measured under 3 seconds", "P2"),
    ]),
    ("CEO Chat", [
        ("Reply latency 1-4s", "POST /api/ceo/chat -> response within 1-4s (Haiku)", "P0"),
        ("Concurrent message lock", "Send 2 messages same session simultaneously -> no interleaving", "P1"),
        ("No framing block leaks", "Reply text never contains [tenant_id=...] prefix", "P0"),
        ("No max-turns errors", "No 'Reached max turns' messages in any reply", "P1"),
        ("Action blocks execute", "create_contact action -> contact appears in CRM", "P0"),
        ("Delegate spawns task", "Delegate block fires background task without 500", "P0"),
        ("Chat history persists", "Refresh page -> prior conversation still shown", "P1"),
    ]),
    ("Agent Delegation (Paperclip)", [
        ("ContentWriter delegation", "Delegate -> inbox row appears within 10-60s", "P0"),
        ("EmailMarketer type", "Delegated email -> row has type=email_sequence", "P0"),
        ("SocialManager drafts", "Delegated social -> row has social_drafts populated", "P0"),
        ("No duplicate rows", "Exactly ONE inbox row per delegation (placeholder + skill curl deduped)", "P0"),
        ("Issues created as todo", "Paperclip issue status=todo, NOT backlog", "P0"),
        ("Wake comment fires run", "No 9-min hangs; --dangerously-skip-permissions in Extra args", "P0"),
        ("CRM auto-enrich", "'email Hanz' -> task description gets Hanz's email appended", "P1"),
        ("Watcher graceful timeout", "After 600s -> watcher exits without 500", "P1"),
    ]),
    ("Inbox", [
        ("Email body populated", "Open email row -> EmailEditor iframe shows body content", "P0"),
        ("Subject readable", "Subject is human text, not raw <html> tag", "P0"),
        ("Subject not 'Untitled'", "Normal agent output produces meaningful subject", "P1"),
        ("Editor buttons render", "Approve/Schedule/Save/Cancel show for draft_pending_approval emails", "P0"),
        ("HTML wrap fallback", "Plain HTML wrapped in branded template; designed HTML untouched", "P1"),
        ("Confirmation rejection", "'Saved to ARIA Inbox' / 'Draft ID:' messages do NOT create rows", "P0"),
        ("Social drafts render", "Per-platform drafts with character counts visible", "P1"),
        ("Edit + save persists", "Edit body -> save -> Supabase row updated", "P0"),
    ]),
    ("Cron / Scheduled Tasks", [
        ("Manual cron run", "POST /api/cron/run-scheduled -> due tasks execute", "P1"),
        ("Scheduler loop healthy", "_scheduler_executor_loop running, no leaked exceptions", "P1"),
        ("Office sync backoff", "Idle period -> 30s interval; poke -> back to 5s", "P2"),
        ("Gmail sync loop", "Runs every 2 min without auth errors", "P1"),
    ]),
    ("System Health", [
        ("Health endpoint", "GET /health -> 200", "P0"),
        ("Paperclip status", "GET /api/paperclip/status -> connected: true", "P0"),
        ("All agents registered", "AGENT_REGISTRY contains all 5 agents", "P0"),
        ("claude.json auto-restore", "Startup log shows restore if backup used", "P1"),
        ("No silent task crashes", "No 'Task exception was never retrieved' warnings", "P1"),
        ("Containers healthy", "docker compose ps -> no Restarting status", "P0"),
    ]),
    ("Deploy Pipeline", [
        ("Webhook fires", "git push -> webhook triggers within 2s", "P0"),
        ("Journal log sequence", "matched -> 200 OK -> executing -> done", "P1"),
        ("No-op deploy fast", "Empty commit deploys in ~4s", "P2"),
        ("Backend edit deploys", "20-40s deploy; new symbol present (grep verifies)", "P0"),
        ("Deploy.sh not duplicated", "Logs show single 'Image aria-backend Built' line", "P1"),
        ("Frontend stays up", "Backend rebuild does NOT take frontend offline", "P0"),
    ]),
    ("Error / Edge Cases", [
        ("Paperclip down fallback", "Stop Paperclip -> CEO chat still replies via local fallback", "P0"),
        ("Supabase down handling", "Supabase outage -> friendly error, no white screen", "P1"),
        ("Invalid tenant_id", "Bad tenant in API -> 404 not 500", "P1"),
        ("Long chat input", "Send >10KB message -> no crash", "P2"),
        ("XSS in chat input", "<script> in input -> sanitized on display", "P0"),
        ("Rate limiting", "Spam /api/ceo/chat -> 429 if limited (or document none)", "P2"),
    ]),
    ("Cross-Browser / Responsive", [
        ("Chrome render", "Dashboard works in Chrome", "P0"),
        ("Firefox render", "Dashboard works in Firefox", "P1"),
        ("Safari render", "Dashboard + email iframe works in Safari", "P1"),
        ("Mobile viewport", "375px width: sidebar collapses, chat usable", "P1"),
    ]),
]

row = 2
counter = 1
for section_name, cases in sections:
    ws.cell(row=row, column=1, value=section_name)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    sc = ws.cell(row=row, column=1)
    sc.font = section_font
    sc.fill = section_fill
    sc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    row += 1
    for title, steps, prio in cases:
        ws.cell(row=row, column=1, value=counter)
        ws.cell(row=row, column=2, value=section_name)
        ws.cell(row=row, column=3, value=title)
        ws.cell(row=row, column=4, value=steps)
        ws.cell(row=row, column=5, value=prio)
        ws.cell(row=row, column=6, value="Not Started")
        ws.cell(row=row, column=7, value="")
        ws.cell(row=row, column=8, value="")
        ws.cell(row=row, column=9, value="")
        for col in range(1, 10):
            c = ws.cell(row=row, column=col)
            c.alignment = wrap
            c.border = border
        row += 1
        counter += 1

# Column widths
widths = [5, 22, 32, 60, 10, 14, 14, 12, 30]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A2"

# Data validation: Status dropdown
from openpyxl.worksheet.datavalidation import DataValidation
dv = DataValidation(type="list", formula1='"Not Started,In Progress,Pass,Fail,Blocked,N/A"', allow_blank=True)
ws.add_data_validation(dv)
dv.add(f"F2:F{row-1}")

prio_dv = DataValidation(type="list", formula1='"P0,P1,P2"', allow_blank=True)
ws.add_data_validation(prio_dv)
prio_dv.add(f"E2:E{row-1}")

out = r"c:\Users\Admin\Documents\ARIA\ARIA_QA_Checklist.xlsx"
wb.save(out)
print(f"Wrote {out}")
print(f"Total cases: {counter-1}")
